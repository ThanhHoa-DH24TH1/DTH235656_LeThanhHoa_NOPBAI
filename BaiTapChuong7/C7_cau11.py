import openpyxl
from openpyxl import Workbook, load_workbook

FILE_NAME = "nhanvien.xlsx"

# ---------------------------------------
# Hàm tạo file Excel nếu chưa tồn tại
# ---------------------------------------
def tao_file_moi():
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"
    ws.append(["STT", "Mã", "Tên", "Tuổi"])
    wb.save(FILE_NAME)
    print("✅ Đã tạo file Excel mới:", FILE_NAME)

# ---------------------------------------
# Hàm lưu nhân viên mới
# ---------------------------------------
def luu_nhanvien(ma, ten, tuoi):
    try:
        wb = load_workbook(FILE_NAME)
        ws = wb.active
    except:
        tao_file_moi()
        wb = load_workbook(FILE_NAME)
        ws = wb.active

    stt = ws.max_row  # STT = số dòng hiện có
    ws.append([stt, ma, ten, tuoi])
    wb.save(FILE_NAME)
    print("💾 Đã lưu nhân viên:", ten)

# ---------------------------------------
# Hàm đọc danh sách nhân viên
# ---------------------------------------
def doc_danhsach():
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    print("\n📋 DANH SÁCH NHÂN VIÊN:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row)

# ---------------------------------------
# Hàm sắp xếp nhân viên theo tuổi tăng dần
# ---------------------------------------
def sapxep_theo_tuoi():
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    # Lấy dữ liệu trừ hàng tiêu đề
    data = list(ws.iter_rows(min_row=2, values_only=True))
    # Sắp xếp theo tuổi (cột thứ 4)
    data.sort(key=lambda x: x[3])

    # Xóa dữ liệu cũ và ghi lại theo thứ tự mới
    ws.delete_rows(2, ws.max_row)
    stt = 1
    for row in data:
        ws.append([stt, row[1], row[2], row[3]])
        stt += 1

    wb.save(FILE_NAME)
    print("✅ Đã sắp xếp nhân viên theo tuổi tăng dần.")

# ---------------------------------------
# Menu chính
# ---------------------------------------
def menu():
    while True:
        print("\n=== QUẢN LÝ NHÂN VIÊN ===")
        print("1. Thêm nhân viên")
        print("2. Xem danh sách")
        print("3. Sắp xếp theo tuổi tăng dần")
        print("0. Thoát")

        chon = input("👉 Chọn chức năng: ")
        if chon == "1":
            ma = input("Nhập mã nhân viên: ")
            ten = input("Nhập tên nhân viên: ")
            tuoi = int(input("Nhập tuổi: "))
            luu_nhanvien(ma, ten, tuoi)
        elif chon == "2":
            doc_danhsach()
        elif chon == "3":
            sapxep_theo_tuoi()
        elif chon == "0":
            print("👋 Thoát chương trình.")
            break
        else:
            print("❌ Lựa chọn không hợp lệ!")

# ---------------------------------------
if __name__ == "__main__":
    menu()
